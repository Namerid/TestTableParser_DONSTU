import openpyxl
import pandas as pd
import pathlib
import shutil
import sys
from pprint import pprint
import re
# import json


# ═══════════════════════════════════════════════════════════════════════════════
# КОНСТАНТЫ — ПУТИ К ПАПКАМ И ФАЙЛАМ
# ═══════════════════════════════════════════════════════════════════════════════

# Корневая рабочая папка, в которой хранятся все промежуточные и итоговые данные.
# Создаётся при запуске и полностью удаляется при повторном запуске.
WORK_FOLDER = "Work_folder"

# Подпапка для хранения подготовленных (скопированных/конвертированных) файлов кафедр
DEPARTMENTS_FOLDER = f"{WORK_FOLDER}\\Departments"

# Подпапка для хранения файлов с баллами студентов (осень и весна)
POINTS_FOLDER = f"{WORK_FOLDER}\\Points"

# Подпапка, в которую будут сохранены итоговые результирующие файлы
OUTPUT_FOLDER = f"{WORK_FOLDER}\\Results"

# Стандартизированные имена файлов баллов внутри рабочей папки.
# Вне зависимости от исходных имён файлы всегда копируются под этими именами.
AUTUMN_POINTS_NAME = "autumn_points.xlsx"
SPRING_POINTS_NAME = "spring_points.xlsx"

# Полные пути к файлам баллов в рабочей папке
AUTUMN_POINTS_PATH = f"{POINTS_FOLDER}\\{AUTUMN_POINTS_NAME}"
SPRING_POINTS_PATH = f"{POINTS_FOLDER}\\{SPRING_POINTS_NAME}"


# ═══════════════════════════════════════════════════════════════════════════════
# КОНСТАНТЫ — СТРУКТУРА ФАЙЛОВ КАФЕДР
# ═══════════════════════════════════════════════════════════════════════════════

# Словарь, который сопоставляет человекочитаемое название каждого столбца
# с его буквенным индексом в Excel-файле кафедры.
# Используется для прямого обращения к ячейкам через openpyxl: sheet["A5"], sheet["BD10"] и т.д.
# Столбцы с именами "Unnamed: N" — это безымянные столбцы, возникшие при экспорте из Excel.
DEPARTMENT_COLUMNS = {
    "Номер": "A",
    "Unnamed: 1": "B",
    "Учебный план": "C",
    "Факультет группы": "D",
    "Блок": "E",
    "Дисциплина, вид учебной работы": "F",
    "Закреплённая кафедра": "G",
    "Курс/Семестр или Курс/Сессия": "H",
    "Группа": "I",
    "Количество студентов": "J",
    "Недель": "K",
    "Вид занятий": "L",
    "Часов (на поток, группу, студента)": "M",
    "Виды контроля": "N",
    "КСР": "O",
    "Индивидуальные занятия": "P",
    "Контрольные": "Q",
    "Оценка по рейтингу": "R",
    "Рефераты": "S",
    "Эссе": "T",
    "РГР": "U",
    "Контрольных работ (заоч)": "V",
    "Консультации (СПО)": "W",
    "Нагрузка, час": "X",
    "Unnamed: 24": "Y",
    "Unnamed: 25": "Z",
    "Unnamed: 26": "AA",
    "Преподаватель": "AB",
    "Unnamed: 28": "AC",
    "Unnamed: 29": "AD",
    "Unnamed: 30": "AE",
    "Номер потока": "AF",
    "Индикатор первой группы потока": "AG",
    "Ауд. рекомендуемая каф.": "AH",
    "Дополнительно часов": "AI",
    "Unnamed: 35": "AJ",
    "Фактически выполнено": "AK",
    "Unnamed: 37": "AL",
    "Время проведения занятий по графику": "AM",
    "Unnamed: 39": "AN",
    "Распределение нагрузки, час": "AO",
    "Unnamed: 41": "AP",
    "Unnamed: 42": "AQ",
    "Доля внебюджет": "AR",
    "Доля иностр.": "AS",
    "Всего/Договор/Иностр": "AT",
    "Уровень образования": "AU",
    "Форма обучения": "AV",
    "Часов на экзамены": "AW",
    "Сам. работа": "AX",
    "Электронные часы": "AY",
    "Нормирующий коэффициент": "AZ",
    "Примечание": "BA",
    "ЗЕТ": "BB",
    "Число вопросов": "BC",
    "Часов в неделю": "BD"
}

# Константы с названиями ключевых столбцов файла кафедры.
# Вынесены отдельно, чтобы использовать их как ключи словаря без риска опечатки.
DEPARTMENT_GROUP_HEADER_NAME = "Группа"
DEPARTMENT_DISCIPLINE_HEADER_NAME = "Дисциплина, вид учебной работы"
DEPARTMENT_NUMBER_OF_STUDENTS_HEADER_NAME = "Количество студентов"
DEPARTMENT_TYPE_OF_ACTIVITY_HEADER_NAME = "Вид занятий"
DEPARTMENT_TEACHER_HEADER_NAME = "Преподаватель"
DEPARTMENT_COURSE_HEADER_NAME = "Курс/Семестр или Курс/Сессия"

# Список столбцов кафедры, которые реально читаются и переносятся в результирующий файл.
# Остальные столбцы из DEPARTMENT_COLUMNS при обработке игнорируются.
DEPARTMENT_COLUMNS_READ_LIST = [
    "Учебный план",
    "Факультет группы",
    "Дисциплина, вид учебной работы",
    "Курс/Семестр или Курс/Сессия",
    "Группа",
    "Количество студентов",
    "Вид занятий",
    "Преподаватель"
]

# Количество строк-заголовков в файле кафедры.
# Данные начинаются со строки DEPARTMENT_HEADER_ROWS_SIZE + 1,
# то есть первые 5 строк — служебная шапка документа, не содержащая записей.
DEPARTMENT_HEADER_ROWS_SIZE = 5


# ═══════════════════════════════════════════════════════════════════════════════
# КОНСТАНТЫ — СТРУКТУРА ФАЙЛОВ БАЛЛОВ
# ═══════════════════════════════════════════════════════════════════════════════

# Словарь, сопоставляющий названия столбцов файла баллов их буквенным индексам.
# Файл баллов содержит данные о прохождении тестов студентами.
# Структура обновлена: ФИО теперь разбито на три отдельных столбца,
# добавлен столбец с названием теста, дисциплина и балл сдвинулись правее.

# Старая структура файла баллов (оставлена для справки)
# POINTS_COLUMNS = {
#     "группа": "A",
#     "ФИО": "B",
#     "дисциплина": "C",
#     "балл": "D",
#     "начало попытки": "E",
#     "конец попытки": "F"
# }

POINTS_COLUMNS = {
    "группа": "A",
    "ФИО": "B",
    "Фамилия": "C",
    "Имя": "D",
    "Отчество": "E",
    "дисциплина": "F",
    "название теста": "G",
    "балл": "H",
    "начало попытки": "I",
    "конец попытки": "J"
}

# Названия ключевых столбцов файла баллов
POINTS_GROUP_HEADER_NAME = "группа"
POINTS_DISCIPLINE_HEADER_NAME = "дисциплина"
POINTS_POINT_HEADER_NAME = "балл"

# Значения, которые считаются «пустыми» и пропускаются при чтении баллов.
# "безоценочно" — особый маркер, означающий, что балл не выставлялся.
POINTS_SKIP_VALUES = [None, "", "безоценочно"]

# Количество строк-заголовков в файле баллов.
# Данные начинаются со второй строки (первая — строка с названиями столбцов).
POINTS_HEADER_ROWS_SIZE = 1


# ═══════════════════════════════════════════════════════════════════════════════
# КОНСТАНТЫ — СТРУКТУРА ВЫХОДНОГО ФАЙЛА
# ═══════════════════════════════════════════════════════════════════════════════

# Словарь, сопоставляющий названия столбцов итогового файла их буквенным индексам.
# Первые 8 столбцов (A–H) переносятся из файла кафедры без изменений.
# Столбцы I и J ("Количество прохождений", "Средний балл") рассчитываются
# по данным из файлов баллов.
OUTPUT_COLUMNS = {
    "Учебный план": "A",
    "Факультет группы": "B",
    "Дисциплина, вид учебной работы": "C",
    "Курс/Семестр или Курс/Сессия": "D",
    "Группа": "E",
    "Количество студентов": "F",
    "Вид занятий": "G",
    "Преподаватель": "H",
    "Количество прохождений": "I",  # Заполняется из файла баллов
    "Средний балл": "J"             # Считается как среднее по всем баллам группы
}

OUTPUT_PASSED_HEADER_NAME = "Количество прохождений"
OUTPUT_AVERAGE_HEADER_NAME = "Средний балл"

# Список столбцов, которые копируются из файла кафедры в выходной файл напрямую
OUTPUT_COLUMNS_WRITE_LIST = [
    "Учебный план",
    "Факультет группы",
    "Дисциплина, вид учебной работы",
    "Курс/Семестр или Курс/Сессия",
    "Группа",
    "Количество студентов",
    "Вид занятий",
    "Преподаватель"
]

# Количество строк-заголовков в выходном файле (одна строка с названиями столбцов)
OUTPUT_HEADER_ROWS_SIZE = 1

# Регулярное выражение для удаления суффиксов из названий дисциплин.
# Удаляет:
#   ", п/г 1"   — обозначение подгруппы (подгруппа 1, 2 и т.д.)
#   ", часть 2" — обозначение части дисциплины (часть 1, 2 и т.д.)
# Оба суффикса опциональны и могут встречаться одновременно.
# "$" гарантирует, что удаляется только хвост строки, а не вхождение внутри.
PATTERN = r"(?:,\s*п/г\s*\d+)?(?:,\s*часть\s*\d+)?$"


# ═══════════════════════════════════════════════════════════════════════════════
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ═══════════════════════════════════════════════════════════════════════════════

# def save_to_json(data, file_path: str | pathlib.Path, indent: int = 2):
#     """
#     Сохраняет произвольные данные (словарь, список и т.д.) в JSON-файл.
#     Используется преимущественно для отладки: позволяет сохранить
#     промежуточный словарь departments_dict и просмотреть его структуру.
#
#     Особенности:
#     - Автоматически создаёт все недостающие родительские директории.
#     - Кириллица сохраняется как есть (ensure_ascii=False), не в виде \uXXXX.
#     - Ключи сортируются по алфавиту для удобного сравнения файлов.
#     - Отступы задаются параметром indent (по умолчанию 2 пробела).
#
#     Параметры:
#         data      — данные для сериализации
#         file_path — путь к выходному файлу (str или pathlib.Path)
#         indent    — количество пробелов для отступа в JSON
#     """
#     file_path = pathlib.Path(file_path)
#     file_path.parent.mkdir(parents=True, exist_ok=True)
#     with open(file_path, 'w', encoding='utf-8') as f:
#         json.dump(
#             data,
#             f,
#             ensure_ascii=False,  # Сохраняем кириллицу напрямую, не экранируем
#             indent=indent,       # Красивое форматирование с отступами
#             sort_keys=True       # Сортировка ключей для удобного чтения
#         )
#     print(f"JSON сохранён: {file_path}")


def to_float(value: str | float | int):
    """
    Безопасно преобразует значение к типу float.
    Поддерживает строки с запятой как десятичным разделителем (например, "3,14").

    Параметры:
        value — строка, число с плавающей точкой или целое число

    Возвращает:
        float или None, если преобразование невозможно
    """
    if isinstance(value, str):
        value = value.replace(",", ".", 1)  # Заменяем только первую запятую
        if value.replace(".", "").isdigit():
            return float(value)
        # Если строка не является числом — возврата нет, функция вернёт None
    elif isinstance(value, float):
        return value
    elif isinstance(value, int):
        return float(value)


def to_int(value: str | int):
    """
    Безопасно преобразует значение к типу int.
    Работает только с целыми числами: дробные строки вроде "3.5" вернут None.

    Параметры:
        value — строка или целое число

    Возвращает:
        int или None, если преобразование невозможно
    """
    if isinstance(value, str) and value.isdigit():
        return int(value)
    elif isinstance(value, int):
        return value


def recreate_work_folder(output_folder_name):
    """
    Пересоздаёт папку по указанному пути: удаляет её вместе со всем содержимым
    (если существует), затем создаёт заново пустой.

    Используется для инициализации рабочих подпапок (Departments, Points)
    перед каждым запуском, чтобы избежать накопления устаревших файлов.

    Параметры:
        output_folder_name — строка или путь к создаваемой папке

    Возвращает:
        pathlib.Path — объект созданной папки
    """
    if pathlib.Path(output_folder_name).exists():
        shutil.rmtree(output_folder_name)  # Удаляем папку рекурсивно

    output_folder = pathlib.Path(output_folder_name)
    output_folder.mkdir(exist_ok=True, parents=True)  # Создаём вместе с родителями

    return output_folder


def remove_suffixes(text: str, pattern: str) -> str:
    """
    Удаляет из строки суффиксы, соответствующие регулярному выражению pattern.

    Применяется для нормализации названий дисциплин: убирает уточнения вида
    ", п/г 1" (подгруппа) и ", часть 2" (часть дисциплины), чтобы привести
    все варианты одной дисциплины к единому виду для последующего сравнения.

    Например:
        "Математика, п/г 1"     → "Математика"
        "Физика, часть 2"       → "Физика"
        "Химия, п/г 1, часть 2" → "Химия"
        "История"               → "История"  (без изменений)

    После удаления суффикса дополнительно обрезаются лишние запятые и пробелы
    в конце строки с помощью второго re.sub.

    Параметры:
        text    — исходная строка с возможным суффиксом
        pattern — регулярное выражение суффиксов для удаления

    Возвращает:
        str — строка без суффиксов, или исходный text, если он пустой/не строка
    """
    if not text or not isinstance(text, str):
        return text  # Защита от None и нестроковых значений

    cleaned = re.sub(pattern, "", text)         # Удаляем суффиксы по шаблону
    cleaned = re.sub(r"[,\s]+$", "", cleaned)   # Убираем хвостовые запятые/пробелы

    return cleaned.strip()                      # Финальная очистка от пробелов по краям


def end_check(text: str, pattern: str) -> bool:
    """
    Проверяет, оканчивается ли строка на подстроку, соответствующую шаблону.

    Используется для определения типа суффикса в названии дисциплины:
    - r"[,\s]*п/г\s*\d+$"    — строка оканчивается на обозначение подгруппы
    - r"часть(?:_к)?\s*\d+$" — строка оканчивается на обозначение части

    Это позволяет принять решение: нужно ли объединять две строки кафедры
    как подгруппы одной и той же дисциплины или как части одной дисциплины.

    Параметры:
        text    — проверяемая строка (полное название дисциплины)
        pattern — регулярный шаблон суффикса

    Возвращает:
        True  — строка оканчивается на данный суффикс
        False — не оканчивается, или text пустой/не строка
    """
    if not text or not isinstance(text, str):
        return False

    return bool(re.search(pattern, text, re.IGNORECASE))


# ═══════════════════════════════════════════════════════════════════════════════
# ФУНКЦИИ ПОДГОТОВКИ ДАННЫХ
# ═══════════════════════════════════════════════════════════════════════════════

def preparation_of_departments(departments_name: str, work_folder_name=WORK_FOLDER, output_folder_name=DEPARTMENTS_FOLDER) -> int:
    """
    Подготавливает файлы кафедр к обработке: копирует .xlsx-файлы
    и конвертирует .xls-файлы в .xlsx, помещая результат в DEPARTMENTS_FOLDER.

    Принимает как папку (обрабатываются все .xls/.xlsx внутри),
    так и одиночный файл.

    Предварительно очищает путь от лишних символов: начальных '&', пробелов
    и кавычек, которые могут появиться при перетаскивании файла в терминал.

    Конвертация .xls → .xlsx выполняется через pandas (engine='xlrd'),
    поскольку openpyxl не поддерживает старый формат .xls.

    Коды возврата:
        0 — успех
        1 — путь не существует
        2 — в папке не найдено ни одного .xls/.xlsx файла
        3 — передан файл неподдерживаемого формата (не .xls и не .xlsx)
        5 — ошибка при копировании или конвертации файла

    Параметры:
        departments_name   — путь к папке или файлу кафедр (введённый пользователем)
        work_folder_name   — путь к корневой рабочей папке (для удаления при ошибке)
        output_folder_name — путь к папке назначения для файлов кафедр
    """
    # Очищаем введённый путь от артефактов терминала (амперсанды, кавычки, пробелы)
    departments_path = pathlib.Path(departments_name.strip("& ").strip("'\""))

    if not departments_path.exists():
        print("Ошибка: Папки/файла не существует")
        shutil.rmtree(work_folder_name)
        return 1

    if departments_path.is_dir():
        # ── Обработка папки ──────────────────────────────────────────────────
        output_folder = recreate_work_folder(output_folder_name)

        xls_files = list(departments_path.glob("*.xls"))    # Старый формат Excel
        xlsx_files = list(departments_path.glob("*.xlsx"))  # Новый формат Excel

        if not xls_files and not xlsx_files:
            print("Ошибка: Файлы не найдены.")
            shutil.rmtree(work_folder_name)
            return 2

        # Конвертируем .xls → .xlsx через pandas + xlrd
        if len(xls_files) > 0:
            print(f"Найдено {len(xls_files)} файл(ов) .xls. Начинаем конвертацию...")
            for xls_file in xls_files:
                file = pd.read_excel(xls_file, engine='xlrd')
                print("Конвертация файла:", xls_file.name)
                try:
                    file.to_excel(
                        f'{str(output_folder)}\\{xls_file.stem}.xlsx',
                        index=False,
                        engine='openpyxl'
                    )
                except Exception as e:
                    print("Ошибка при конвертации файла:", xls_file.name)
                    print("Подробности:", e)
                    # Ошибка логируется, но обработка остальных файлов продолжается

        # Просто копируем .xlsx-файлы без преобразования
        if len(xlsx_files) > 0:
            print(f"\nНайдено {len(xlsx_files)} файл(ов) .xlsx. Начинаем копирование...")
            for xlsx_file in xlsx_files:
                print("Копирование файла:", xlsx_file.name)
                try:
                    shutil.copy(xlsx_file, f'{str(output_folder)}\\{xlsx_file.name}')
                except Exception as e:
                    print("Ошибка при копировании файла:", xlsx_file.name)
                    print("Подробности:", e)

    elif departments_path.is_file():
        # ── Обработка одиночного файла ───────────────────────────────────────
        output_folder = recreate_work_folder(output_folder_name)

        if departments_path.suffix.lower() == ".xls":
            # Конвертируем одиночный .xls-файл в .xlsx
            file = pd.read_excel(departments_path, engine='xlrd')
            print("Конвертация файла:", departments_path.name)
            try:
                file.to_excel(
                    f'{str(output_folder)}\\{departments_path.stem}.xlsx',
                    index=False,
                    engine='openpyxl'
                )
            except Exception as e:
                print("Ошибка при конвертации файла:", departments_path.name)
                print("Подробности:", e)
                shutil.rmtree(work_folder_name)
                return 5

        elif departments_path.suffix.lower() == ".xlsx":
            # Копируем одиночный .xlsx-файл без преобразования
            print("Копирование файла:", departments_path.name)
            try:
                shutil.copy(departments_path, f'{str(output_folder)}\\{departments_path.name}')
            except Exception as e:
                print("Ошибка при копировании файла:", departments_path.name)
                print("Подробности:", e)
                shutil.rmtree(work_folder_name)
                return 5
        else:
            print("Ошибка: Неверный тип файла.")
            shutil.rmtree(work_folder_name)
            return 3

    return 0


def preparation_of_points(autumn_points_name: str, spring_points_name: str, work_folder_name=WORK_FOLDER, output_folder_name=POINTS_FOLDER) -> int:
    """
    Подготавливает файлы баллов к обработке: копирует осенний и весенний
    файлы в рабочую папку POINTS_FOLDER под стандартизированными именами.

    Принимает только .xlsx-файлы. Если передан файл другого формата — ошибка.
    Каждый из двух файлов можно пропустить, оставив поле ввода пустым.
    Однако хотя бы один файл должен быть указан — иначе обрабатывать нечего.

    Перед копированием проверяет:
    1. Существование родительской рабочей папки (WORK_FOLDER).
    2. Существование указанных файлов-источников.
    3. Расширение указанных файлов (.xlsx).

    Коды возврата:
        0 — успех
        1 — один из указанных файлов не существует
        3 — неверный тип файла (не .xlsx)
        4 — родительская рабочая папка не существует
        5 — ошибка при копировании файла
        6 — не указан ни один файл баллов

    Параметры:
        autumn_points_name — путь к файлу осенних баллов (может быть пустым)
        spring_points_name — путь к файлу весенних баллов (может быть пустым)
        work_folder_name   — корневая рабочая папка (для удаления при ошибке)
        output_folder_name — папка назначения для файлов баллов
    """
    # Флаги: пропущен ли каждый из файлов (пользователь оставил поле пустым)
    autumn_skip = True
    spring_skip = True

    output_folder_path = pathlib.Path(output_folder_name)

    # Проверяем, что родительская рабочая папка уже создана (должна быть создана ранее)
    if not output_folder_path.parent.exists():
        print("Ошибка: Рабочая папка не существует")
        shutil.rmtree(work_folder_name)
        return 4

    output_folder_path.mkdir(exist_ok=True)  # Создаём подпапку Points, если её нет

    if not autumn_points_name.isspace() and autumn_points_name:
        # Пользователь указал осенний файл — обрабатываем его
        autumn_skip = False

        # Очищаем путь от артефактов терминала (амперсанды, кавычки, пробелы)
        autumn_points_path = pathlib.Path(autumn_points_name.strip("& ").strip("'\""))

        if not autumn_points_path.exists():
            print("Ошибка: Файла осенних баллов не существует")
            shutil.rmtree(work_folder_name)
            return 1

        if autumn_points_path.suffix.lower() == ".xlsx":
            # Копируем файл осенних баллов под стандартным именем autumn_points.xlsx
            print("Копирование файла:", autumn_points_path.name)
            try:
                shutil.copy(autumn_points_path, f'{str(output_folder_path)}\\{AUTUMN_POINTS_NAME}')
            except Exception as e:
                print("Ошибка при копировании файла:", autumn_points_path.name)
                print("Подробности:", e)
                shutil.rmtree(work_folder_name)
                return 5
        else:
            print("Ошибка: Неверный тип файла:", autumn_points_path.name)
            shutil.rmtree(work_folder_name)
            return 3

    if not spring_points_name.isspace() and spring_points_name:
        # Пользователь указал весенний файл — обрабатываем его
        spring_skip = False

        # Очищаем путь от артефактов терминала (амперсанды, кавычки, пробелы)
        spring_points_path = pathlib.Path(spring_points_name.strip("& ").strip("'\""))

        if not spring_points_path.exists():
            print("Ошибка: Файла весенних баллов не существует")
            shutil.rmtree(work_folder_name)
            return 1

        if spring_points_path.suffix.lower() == ".xlsx":
            # Копируем файл весенних баллов под стандартным именем spring_points.xlsx
            print("Копирование файла:", spring_points_path.name)
            try:
                shutil.copy(spring_points_path, f'{str(output_folder_path)}\\{SPRING_POINTS_NAME}')
            except Exception as e:
                print("Ошибка при копировании файла:", spring_points_path.name)
                print("Подробности:", e)
                shutil.rmtree(work_folder_name)
                return 5
        else:
            print("Ошибка: Неверный тип файла:", spring_points_path.name)
            shutil.rmtree(work_folder_name)
            return 3

    # Оба файла пропущены — дальнейшая обработка бессмысленна
    if autumn_skip and spring_skip:
        print("Ошибка: Не указано ни одного файла баллов для обработки.")
        shutil.rmtree(work_folder_name)
        return 6

    return 0


# ═══════════════════════════════════════════════════════════════════════════════
# ФУНКЦИИ ЧТЕНИЯ ДАННЫХ
# ═══════════════════════════════════════════════════════════════════════════════

def read_department_file(
    file_path: pathlib.Path,
    department_header_rows_size=DEPARTMENT_HEADER_ROWS_SIZE,
    department_columns=DEPARTMENT_COLUMNS,
    department_teacher_header_name=DEPARTMENT_TEACHER_HEADER_NAME,
    department_number_of_students_header_name=DEPARTMENT_NUMBER_OF_STUDENTS_HEADER_NAME,
    department_group_header_name=DEPARTMENT_GROUP_HEADER_NAME,
    department_discipline_header_name=DEPARTMENT_DISCIPLINE_HEADER_NAME,
    department_course_header_name=DEPARTMENT_COURSE_HEADER_NAME,
    department_type_of_activity_header_name=DEPARTMENT_TYPE_OF_ACTIVITY_HEADER_NAME,
    department_columns_read_list=DEPARTMENT_COLUMNS_READ_LIST,
    pattern=PATTERN
):
    """
    Читает файл кафедры и возвращает структурированный словарь данных.

    Возвращаемая структура:
        {
            "ИТ-21": [
                {
                    "Учебный план": "...",
                    "Факультет группы": "...",
                    "Дисциплина, вид учебной работы": "Математика",
                    "Полное название дисциплины": "Математика, п/г 1",
                    "Курс/Семестр или Курс/Сессия": "1/1",
                    "Группа": "ИТ-21",
                    "Количество студентов": "25",
                    "Вид занятий": "Лекция, Практика",
                    "Преподаватель": "Иванов И.И., Петров П.П."
                },
                ...
            ],
            ...
        }

    Ключ словаря — название группы. Значение — список записей (строк) по этой группе.

    Логика обработки строк:
    1. Строки с пустой группой или дисциплиной пропускаются.
    2. Из названия дисциплины удаляются суффиксы ", п/г N" и ", часть N"
       для нормализации (результат сохраняется в поле дисциплины).
       Полное оригинальное название хранится в "Полное название дисциплины".
    3. У поля "Преподаватель" обрезается первый символ (служебный артефакт Excel).
    4. Каждая новая строка проверяется на возможность объединения
       с уже добавленными записями той же группы:
         - Если обе строки — подгруппы одной дисциплины (суффикс "п/г"):
           → суммируется количество студентов, объединяются преподаватели.
         - Если обе строки — части одной дисциплины (суффикс "часть"):
           → только объединяются преподаватели.
         - Иначе — добавляется новая отдельная запись.

    Финальная дедупликация:
    После первичного чтения выполняется второй проход: строки с одинаковой
    дисциплиной, курсом и группой (но разными видами занятий) объединяются —
    склеиваются виды занятий и преподаватели через запятую, выбирается
    максимальное количество студентов. Дубликат удаляется из списка.

    Параметры:
        file_path                                  — путь к файлу кафедры (.xlsx)
        department_header_rows_size                — число строк-заголовков
        department_columns                         — маппинг названий → буквы столбцов
        department_teacher_header_name             — название столбца с преподавателем
        department_number_of_students_header_name  — название столбца с числом студентов
        department_group_header_name               — название столбца с группой
        department_discipline_header_name          — название столбца с дисциплиной
        department_course_header_name              — название столбца с курсом/семестром
        department_type_of_activity_header_name    — название столбца с видом занятий
        department_columns_read_list               — список считываемых столбцов
        pattern                                    — регулярное выражение для удаления суффиксов

    Возвращает:
        dict — словарь { группа: [записи] }, или None при ошибке открытия файла
    """
    departments_dict = {}

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print("Ошибка при открытии файла:", file_path.name)
        print("Подробности:", e)
        return None

    sheet = wb.active
    n = sheet.max_row

    # ── Первичное чтение строк ───────────────────────────────────────────────
    for i in range(department_header_rows_size + 1, n):

        # Пропускаем строки без группы или дисциплины (пустые/объединённые ячейки)
        if (sheet[department_columns[department_group_header_name] + str(i)].value in [None, ""] or
                sheet[department_columns[department_discipline_header_name] + str(i)].value in [None, ""]):
            continue

        group = str(sheet[department_columns[department_group_header_name] + str(i)].value)

        # Нормализованное название дисциплины (без суффиксов п/г и часть)
        discipline = remove_suffixes(
            sheet[department_columns[department_discipline_header_name] + str(i)].value, pattern
        )
        # Полное оригинальное название (нужно для определения суффикса при объединении)
        discipline_full_name = str(
            sheet[department_columns[department_discipline_header_name] + str(i)].value
        )

        # Первый символ в поле преподавателя — служебный, обрезаем его
        teacher = str(sheet[department_columns[department_teacher_header_name] + str(i)].value)
        if teacher not in ["None", ""]:
            teacher = teacher[1:]

        add_check = False  # Флаг: была ли текущая строка объединена с существующей записью

        if group not in departments_dict:
            departments_dict[group] = []

        # Формируем словарь для текущей строки Excel
        add_dict = {}
        for column in department_columns_read_list:
            if column == department_discipline_header_name:
                # Сохраняем и нормализованное, и полное название
                add_dict[column] = discipline
                add_dict["Полное название дисциплины"] = discipline_full_name
            elif column == department_teacher_header_name:
                add_dict[column] = teacher
            elif (column == department_type_of_activity_header_name and
                  sheet[department_columns[column] + str(i)].value is None):
                # Если вид занятий не указан — ставим пустую строку вместо "None"
                add_dict[column] = ""
            else:
                add_dict[column] = str(sheet[department_columns[column] + str(i)].value)

        # ── Попытка объединить с существующей записью ────────────────────────
        for record in departments_dict[group]:

            # Проверяем: обе строки — подгруппы (суффикс "п/г") одной дисциплины?
            is_pg_current = end_check(discipline_full_name, r"[,\s]*п/г\s*\d+$")
            is_pg_record = end_check(record["Полное название дисциплины"], r"[,\s]*п/г\s*\d+$")
            both_are_pg = (is_pg_current == is_pg_record) and is_pg_current

            # Проверяем: обе строки — части (суффикс "часть") одной дисциплины?
            is_part_current = end_check(discipline_full_name, r"часть(?:_к)?\s*\d+$")
            is_part_record = end_check(record["Полное название дисциплины"], r"часть(?:_к)?\s*\d+$")
            both_are_part = (is_part_current == is_part_record) and is_part_current

            # Проверяем совпадение контекста: курс, группа, вид занятий
            same_course = add_dict[department_course_header_name] == record[department_course_header_name]
            same_group = group == record[department_group_header_name]
            activity_matches = add_dict[department_type_of_activity_header_name] in record[department_type_of_activity_header_name]

            if both_are_pg and same_course and same_group and activity_matches:
                # Подгруппы: суммируем студентов и добавляем преподавателя
                total_students = to_int(record[department_number_of_students_header_name])
                total_students += to_int(add_dict[department_number_of_students_header_name])
                record[department_number_of_students_header_name] = str(total_students)
                if teacher not in record[department_teacher_header_name]:
                    record[department_teacher_header_name] += f", {teacher}"
                add_check = True
                break

            elif both_are_part and same_course and same_group and activity_matches:
                # Части дисциплины: только добавляем преподавателя, студентов не суммируем
                if teacher not in record[department_teacher_header_name]:
                    record[department_teacher_header_name] += f", {teacher}"
                add_check = True
                break

        # Если строку не объединили — добавляем как новую запись
        if not add_check:
            departments_dict[group].append(add_dict)

    # ── Финальная дедупликация ───────────────────────────────────────────────
    # Второй проход: ищем записи с одинаковой дисциплиной + курсом + группой,
    # которые не были объединены в первом проходе (например, разные виды занятий).
    # Объединяем их: склеиваем виды занятий и преподавателей, берём макс. студентов.
    for group in departments_dict:
        i = 0
        for i in range(len(departments_dict[group])):
            if i >= len(departments_dict[group]):
                break
            for j in range(len(departments_dict[group])):
                if j >= len(departments_dict[group]) or i >= len(departments_dict[group]):
                    break
                if j != i:
                    rec_i = departments_dict[group][i]
                    rec_j = departments_dict[group][j]

                    # Дубликат: та же дисциплина, тот же курс, та же группа
                    if (rec_i[department_discipline_header_name] == rec_j[department_discipline_header_name] and
                            rec_i[department_course_header_name] == rec_j[department_course_header_name] and
                            rec_i[department_group_header_name] == rec_j[department_group_header_name]):

                        # Добавляем преподавателя из дубликата, если его ещё нет
                        if rec_j[department_teacher_header_name] not in rec_i[department_teacher_header_name]:
                            rec_i[department_teacher_header_name] += f", {rec_j[department_teacher_header_name]}"

                        # Добавляем вид занятий из дубликата, если его ещё нет
                        if rec_j[department_type_of_activity_header_name] not in rec_i[department_type_of_activity_header_name]:
                            rec_i[department_type_of_activity_header_name] += f", {rec_j[department_type_of_activity_header_name]}"

                        # Оставляем максимальное число студентов из двух записей
                        i_students = to_int(rec_i[department_number_of_students_header_name])
                        j_students = to_int(rec_j[department_number_of_students_header_name])
                        if i_students < j_students:
                            rec_i[department_number_of_students_header_name] = str(j_students)

                        # Удаляем поглощённый дубликат
                        departments_dict[group].pop(j)

    return departments_dict


def read_point_file(
    file_path: pathlib.Path,
    point_header_rows_size=POINTS_HEADER_ROWS_SIZE,
    points_columns=POINTS_COLUMNS,
    points_group_header_name=POINTS_GROUP_HEADER_NAME,
    points_discipline_header_name=POINTS_DISCIPLINE_HEADER_NAME,
    points_point_header_name=POINTS_POINT_HEADER_NAME,
    points_skip_values=POINTS_SKIP_VALUES
):
    """
    Читает файл баллов и возвращает словарь с результатами тестирования.

    Возвращаемая структура:
        {
            "ИТ-21": {
                "Математика": [85.0, 72.0, 91.0, ...],
                "Физика":     [60.0, 78.0, ...],
                ...
            },
            "ИТ-22": { ... },
            ...
        }

    Ключ первого уровня — название группы.
    Ключ второго уровня — название дисциплины.
    Значение — список всех баллов студентов группы по данной дисциплине.

    Строки со значениями из POINTS_SKIP_VALUES (None, "", "безоценочно")
    в любом из трёх столбцов (группа, дисциплина, балл) пропускаются.

    Каждый балл преобразуется в float через to_float() для единообразия.

    Коды возврата при ошибке:
        1 — файл не существует (штатная ситуация: файл был пропущен пользователем)
        2 — файл существует, но не удалось его открыть

    Параметры:
        file_path                     — путь к файлу баллов (.xlsx)
        point_header_rows_size        — число строк-заголовков в файле
        points_columns                — маппинг названий столбцов → буквы Excel
        points_group_header_name      — название столбца с группой
        points_discipline_header_name — название столбца с дисциплиной
        points_point_header_name      — название столбца с баллом
        points_skip_values            — список значений, которые нужно пропустить

    Возвращает:
        dict — словарь { группа: { дисциплина: [баллы] } }
        1    — если файл не существует
        2    — если файл не удалось открыть
    """
    points_dict = {}

    # Файл может не существовать, если пользователь пропустил его при вводе — это нормально
    if not file_path.exists():
        return 1

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print("Ошибка при открытии файла:", file_path.name)
        print("Подробности:", e)
        return 2

    sheet = wb.active
    n = sheet.max_row

    for i in range(point_header_rows_size + 1, n):
        try:
            group = sheet[points_columns[points_group_header_name] + str(i)].value
            discipline = sheet[points_columns[points_discipline_header_name] + str(i)].value

            # Убираем лишние пробелы — иначе "Математика " и "Математика" не совпадут при поиске
            if type(discipline) == str:
                discipline = discipline.strip()

            points = sheet[points_columns[points_point_header_name] + str(i)].value

            # Пропускаем строки, где группа, дисциплина или балл отсутствуют/недопустимы
            if not (group in points_skip_values or
                    discipline in points_skip_values or
                    points in points_skip_values):

                if group not in points_dict:
                    points_dict[group] = {}

                if discipline not in points_dict[group]:
                    points_dict[group][discipline] = []

                # Добавляем балл в список, приводя к float
                points_dict[group][discipline].append(to_float(points))

        except Exception as e:
            # Ошибка в строке не прерывает обработку — пропускаем и идём дальше
            print("Ошибка при чтении баллов из файла:", file_path.name)
            print("Подробности:", e)
            continue

    return points_dict


# ═══════════════════════════════════════════════════════════════════════════════
# ОСНОВНАЯ ФУНКЦИЯ ОБРАБОТКИ
# ═══════════════════════════════════════════════════════════════════════════════

def processing(
    departments_folder=DEPARTMENTS_FOLDER,
    autumn_points_path=AUTUMN_POINTS_PATH,
    spring_points_path=SPRING_POINTS_PATH,
    points_folder=POINTS_FOLDER,
    output_folder=OUTPUT_FOLDER,
    output_columns=OUTPUT_COLUMNS,
    output_header_rows_size=OUTPUT_HEADER_ROWS_SIZE,
    output_columns_write_list=OUTPUT_COLUMNS_WRITE_LIST,
    output_passed_header_name=OUTPUT_PASSED_HEADER_NAME,
    output_average_header_name=OUTPUT_AVERAGE_HEADER_NAME,
    department_course_header_name=DEPARTMENT_COURSE_HEADER_NAME,
    department_discipline_header_name=DEPARTMENT_DISCIPLINE_HEADER_NAME
):
    """
    Основная функция обработки данных. Объединяет данные кафедр и баллов,
    формирует итоговые Excel-файлы с результатами тестирования.

    Алгоритм работы:
    1. Читает файлы осенних и весенних баллов в память (read_point_file).
       Если файл был пропущен пользователем — помечает его флагом skip.
    2. Для каждого .xlsx-файла из папки DEPARTMENTS_FOLDER:
       а) Читает данные кафедры (read_department_file).
       б) Создаёт новый Excel-файл с заголовками из OUTPUT_COLUMNS.
       в) Для каждой записи (группа + дисциплина + курс) определяет семестр:
          - Чётный номер семестра/курса → весенние баллы (spring_points).
          - Нечётный номер → осенние баллы (autumn_points).
       г) Ищет баллы по ключам [группа][дисциплина] в соответствующем словаре.
       д) Если баллы найдены — записывает количество прохождений и средний балл.
       е) Применяет форматирование: жирный заголовок, тонкие границы для всех ячеек.
       ж) Сохраняет файл в OUTPUT_FOLDER с тем же именем, что и файл кафедры.
    3. Возвращает 0 при успехе.

    Определение семестра:
        Берётся последний символ строки "Курс/Семестр или Курс/Сессия".
        Если это цифра:
            чётная → весна (семестры 2, 4, 6, 8)
            нечётная → осень (семестры 1, 3, 5, 7)

    Параметры:
        departments_folder         — папка с файлами кафедр
        autumn_points_path         — путь к файлу осенних баллов
        spring_points_path         — путь к файлу весенних баллов
        points_folder              — папка с файлами баллов (не используется напрямую)
        output_folder              — папка для сохранения результатов
        output_columns             — маппинг столбцов выходного файла
        output_header_rows_size    — число строк-заголовков в выходном файле
        output_columns_write_list  — список столбцов, переносимых из файла кафедры
        output_passed_header_name  — название столбца "Количество прохождений"
        output_average_header_name — название столбца "Средний балл"
        department_course_header_name     — название столбца с курсом/семестром
        department_discipline_header_name — название столбца с дисциплиной

    Возвращает:
        0 при успехе, None при ошибке создания выходной папки
    """
    print("\nОсновная обработка данных начата...")

    autumn_points_path = pathlib.Path(autumn_points_path)
    spring_points_path = pathlib.Path(spring_points_path)
    departments_path = pathlib.Path(departments_folder)
    output_folder_path = pathlib.Path(output_folder)

    # Флаги: нужно ли пропустить тот или иной файл баллов
    autumn_points_skip = False
    spring_points_skip = False

    # Загружаем оба файла баллов целиком в память — они используются для всех кафедр.
    # Если файл не существует (код возврата 1) — помечаем его как пропущенный.
    autumn_points = read_point_file(autumn_points_path)
    if autumn_points == 1:
        autumn_points_skip = True
    else:
        print("Чтение осенних баллов")

    spring_points = read_point_file(spring_points_path)
    if spring_points == 1:
        spring_points_skip = True
    else:
        print("Чтение весенних баллов\n")

    # Пересоздаём выходную папку (удаляем старые результаты)
    try:
        if output_folder_path.exists():
            shutil.rmtree(output_folder_path)
        output_folder_path.mkdir(exist_ok=True)
    except Exception as e:
        print("Ошибка при создании выходной папки:", output_folder)
        print("Подробности:", e)
        return None  # Сигнал об ошибке

    # ── Обработка каждого файла кафедры ─────────────────────────────────────
    for department in departments_path.glob("*.xlsx"):
        print("Обработка файла кафедры:", department.name)

        try:
            # Читаем данные кафедры: { группа: [записи] }
            departments_dict = read_department_file(department)

            # Создаём новый Excel-файл для результатов
            wb = openpyxl.Workbook()
            sheet = wb.active

            # Записываем строку заголовков в первую строку
            for header_name in output_columns:
                sheet[output_columns[header_name] + "1"] = header_name

            row_index = output_header_rows_size + 1  # Начинаем запись со второй строки
            discipline = ""

            for group in departments_dict:
                # Получаем словари баллов для текущей группы (None, если группы нет в файле баллов)
                if not autumn_points_skip:
                    autumn_points_group = autumn_points.get(group)
                if not spring_points_skip:
                    spring_points_group = spring_points.get(group)

                for record in departments_dict[group]:
                    # Переносим данные из записи кафедры в выходной файл
                    for column in output_columns_write_list:
                        if column == department_discipline_header_name:
                            discipline = record[column]  # Запоминаем текущую дисциплину
                        sheet[output_columns[column] + str(row_index)] = record[column]

                    course = record[department_course_header_name]

                    # Определяем семестр и ищем баллы только если есть курс и дисциплина
                    if len(course) > 0 and discipline != "":
                        if course[-1].isdigit() and course[-1] != 0:
                            course_num = int(course[-1])

                            if course_num % 2 == 0 and not spring_points_skip:
                                # Чётный семестр → ищем в весенних баллах
                                if spring_points_group is not None:
                                    points = spring_points_group.get(discipline)
                                    if points is not None:
                                        sheet[output_columns[output_passed_header_name] + str(row_index)] = str(len(points))
                                        sheet[output_columns[output_average_header_name] + str(row_index)] = f"{sum(points) / len(points):.2f}"
                            elif not autumn_points_skip:
                                # Нечётный семестр → ищем в осенних баллах
                                if autumn_points_group is not None:
                                    points = autumn_points_group.get(discipline)
                                    if points is not None:
                                        sheet[output_columns[output_passed_header_name] + str(row_index)] = str(len(points))
                                        sheet[output_columns[output_average_header_name] + str(row_index)] = f"{sum(points) / len(points):.2f}"

                    row_index += 1

            # ── Форматирование выходного файла ───────────────────────────────
            bold_font = openpyxl.styles.Font(bold=True)
            thin = openpyxl.styles.Side(border_style="thin", color="000000")
            border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)

            # Жирный шрифт для всей первой строки (заголовки)
            for cell in sheet[1]:
                cell.font = bold_font

            # Тонкие границы для всех заполненных ячеек таблицы
            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row,
                min_col=1, max_col=sheet.max_column
            ):
                for cell in row:
                    cell.border = border

            # Сохраняем результирующий файл с тем же именем, что и исходный файл кафедры
            wb.save(f"{output_folder}\\{department.name}")

        except Exception as e:
            # Ошибка в одном файле не прерывает обработку остальных
            print("Ошибка при обработке файла кафедры:", department.name)
            print("Подробности:", e)
            continue

    return 0


# ═══════════════════════════════════════════════════════════════════════════════
# ТОЧКА ВХОДА
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    """
    Точка входа программы. Управляет последовательным выполнением трёх этапов:
      1. preparation_of_departments — подготовка файлов кафедр.
      2. preparation_of_points      — подготовка файлов баллов.
      3. processing                 — основная обработка и формирование результатов.

    Если любой из этапов завершается с ненулевым кодом возврата —
    выводится сообщение об ошибке и программа завершается.

    Перехватывает KeyboardInterrupt (Ctrl+C) и любые непредвиденные исключения
    для корректного завершения вместо вывода трейсбека Python.

    Все этапы запрашивают пути к файлам у пользователя через input().
    После завершения (успешного или с ошибкой) программа ждёт нажатия клавиши,
    чтобы окно терминала не закрылось автоматически (актуально при запуске
    двойным кликом в Windows).
    """
    try:
        # Этап 1: подготовка файлов кафедр
        if preparation_of_departments(input("Введите путь к папке/файлу кафедр: ")) != 0:
            input("\nОшибка при подготовке файлов кафедр. Завершение работы. Нажмите любую клавишу для выхода...")
            sys.exit()

        # Этап 2: подготовка файлов баллов
        # Каждый из двух файлов можно пропустить, нажав Enter
        if preparation_of_points(
            input("\nВведите путь к файлу осенних баллов (нажмите Enter, чтобы пропустить): "),
            input("Введите путь к файлу весенних баллов (нажмите Enter, чтобы пропустить): ")
        ) != 0:
            input("\nОшибка при подготовке файлов баллов. Завершение работы. Нажмите любую клавишу для выхода...")
            sys.exit()

        # Этап 3: основная обработка
        if processing() != 0:
            input("\nОшибка при обработке данных. Завершение работы. Нажмите любую клавишу для выхода...")
            sys.exit()
        else:
            input("\nОбработка завершена. Нажмите любую клавишу для выхода...")

    except KeyboardInterrupt:
        # Пользователь прервал выполнение через Ctrl+C
        input("\nРабота прервана пользователем. Завершение работы. Нажмите любую клавишу для выхода...")
        sys.exit()
    except Exception as e:
        # Ловим любые другие неожиданные ошибки
        print("\nОшибка:", e)
        input("Завершение работы. Нажмите любую клавишу для выхода...")
        sys.exit()


if __name__ == "__main__":
    main()